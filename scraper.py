from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import openpyxl
import json
with open("config.json","r") as f:
    config=json.loads(f.read())
max_reviews_per_branch = config["max_reviews_per_branch"]
max_branch = config["max_branch"]
search_phrase = config["search_phrase"]


def element_wait(driver , method , element , timeout=10): 
    return WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((method,element)))


def which_is_first(browser , method1 , method2 ,  input1 , input2):
    while True:
        time.sleep(0.1)
        try:
            browser.find_element(method1 , input1)
        except:
            try:
                browser.find_element(method2 , input2)
            except:
                pass
            else:
                return 2
        else:
            return 1


def extract_single_branch(browser):
        
    Reviews_tab = element_wait(browser,"xpath","//div[@class='Gpq6kf fontTitleSmall'][contains(text(),'Reviews')]")
    Reviews_tab.click()
    time.sleep(0.5)
    element_wait(browser , "xpath","//div[@class='m3rned']").click()
    action = ActionChains(browser)
    action.send_keys(Keys.END)
    action.perform()
    if not browser.find_elements("css selector","div.qjESne"):
        try:element_wait(browser , "xpath","//span[contains(text(),'More reviews')]").click()
        except:pass
    element_wait(browser , "css selector","div.jftiEf.fontBodyMedium")
    reviews = browser.find_elements("css selector","div.jftiEf.fontBodyMedium")
    while len(reviews) < max_reviews_per_branch and browser.find_elements("css selector","div.qjESne"):

        reviews = browser.find_elements("css selector","div.jftiEf.fontBodyMedium")
        
        action.send_keys(Keys.HOME)
        action.perform()
        time.sleep(1)
        action.send_keys(Keys.END)
        action.perform()
        time.sleep(1)
    reviews = reviews[0:max_reviews_per_branch]
    review_list=[]
    for review in reviews:
        review_dict = {}
        review_dict["name"] = review.find_element("css selector" ,"div.d4r55").text
        review_dict["text"] = review.find_element("css selector" , "div.MyEned").text
        review_dict["rating"] = len(review.find_elements("css selector","img.hCCjke.vzX5Ic"))
        image_urls = []
        images = review.find_elements("css selector" , "div.KtCyie button.Tya61d")
        for image in images:
            matches = re.findall(r"http.*no",image.get_attribute("style"))
            if matches:
                image_urls.append(matches[0])
        review_dict["images"] = str(image_urls)
        review_list.append(review_dict)
    return review_list


def extract_multi_branch(browser):
    branches = browser.find_elements("css selector","a.hfpxzc")
    while len(branches) < max_branch and not browser.find_elements("xpath","//span[contains(text(),'reached the end')]"):

        branches = browser.find_elements("css selector","a.hfpxzc")
        branches[0].click()
        action = ActionChains(browser)
        action.send_keys(Keys.HOME)
        action.perform()
        time.sleep(1)
        action.send_keys(Keys.END)
        action.perform()
        time.sleep(1)

    branches = branches[0:max_branch]
    branch_list = []
    for branch in branches:
        browser.execute_script("arguments[0].scrollIntoView();", branch)
        branch.click()
        time.sleep(0.5)
        branch_dict ={}
        branch_dict["address"]=""
        try:branch_dict["address"] = element_wait(browser,'css selector','button[data-item-id="address"]').text
        except:pass
        branch_dict["reviews"] = extract_single_branch(browser)
        close_btn = element_wait(browser,'css selector',"button[data-disable-idom='true'][aria-label='Close']")
        close_btn.click()
        branch_list.append(branch_dict)
    return branch_list
    

thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))


def bordering_range(sheet,From, To, border):
    for i in range(From[0], To[0] + 1):
        for j in range(From[1], To[1] + 1):
            sheet.cell(row=i, column=j).border = border

def merge_and_border(sheet,From, To):
    for i in range(From[0], To[0] + 1):
        for j in range(From[1], To[1] + 1):
            sheet.cell(row=i, column=j).border = thin_border
    sheet.merge_cells(start_row=From[0], start_column=From[1], end_row=To[0], end_column=To[1])


browser = webdriver.Firefox()
browser.maximize_window()
browser.get("https://google.com/maps")
search_box = browser.find_element("id","searchboxinput")
search_box.send_keys(search_phrase)
search_box.send_keys(Keys.ENTER)
element_wait(browser,'css selector','div[role="main"]')

page_type = which_is_first(browser,"css selector","css selector",':not([aria-label])[role="main"]','[aria-label][role="main"]')

if page_type == 1 : # multi branch
    branch_list = extract_multi_branch(browser)

elif page_type == 2 : #single branch
    branch_dict ={}
    branch_dict["address"] = element_wait(browser,'css selector','button[data-item-id="address"]').text
    branch_dict["reviews"]  = extract_single_branch(browser)
    branch_list = [branch_dict]

browser.close()
workbook = openpyxl.Workbook()
sheet = workbook.active

coll_i=0
for branch in branch_list:

    
    sheet.cell(row=1, column=5*coll_i+1 ).value=branch["address"]
    merge_and_border(sheet,(1,coll_i*5+1),(1,coll_i*5+5))
    sheet.cell(row=2, column=5*coll_i+1 ).value="index"
    sheet.cell(row=2, column=5*coll_i+2 ).value="name"
    sheet.cell(row=2, column=5*coll_i+3 ).value="text"
    sheet.cell(row=2, column=5*coll_i+4 ).value="rating"
    sheet.cell(row=2, column=5*coll_i+5 ).value="images"

    row_i = 0
    for review in branch["reviews"]:
        sheet.cell(row=row_i+3, column=5*coll_i+row_i+1 ).value= row_i +1
        sheet.cell(row=row_i+3, column=5*coll_i+row_i+2 ).value= review["name"]
        sheet.cell(row=row_i+3, column=5*coll_i+row_i+3 ).value= review["text"]
        sheet.cell(row=row_i+3, column=5*coll_i+row_i+4 ).value= review["rating"]
        sheet.cell(row=row_i+3, column=5*coll_i+row_i+5 ).value= review["images"]
        row_i +=1
    bordering_range(sheet,(2,coll_i+1),(row_i+3,coll_i+5),thin_border)
    coll_i += 1

workbook.save("./"+search_phrase+".xlsx")


